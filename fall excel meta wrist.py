import numpy as np
import torch
import pandas as pd
import os
from sklearn.preprocessing import MinMaxScaler
from denoising_diffusion_pytorch import Unet1D, GaussianDiffusion1D, Trainer1D, Dataset1D
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Directory containing the Excel files
excel_dir = 'Meta Wrist/'

# Get a list of all Excel files in the directory
excel_files = [file for file in os.listdir(excel_dir) if file.endswith('.xlsx')]

# Split the data into trials based on empty spaces
trials = []
current_trial = []

for excel_file in excel_files:
    # Construct the full path to the Excel file
    excel_file_path = os.path.join(excel_dir, excel_file)

    try:
        # Open the Excel file and read sheet names
        excel_workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        sheet_names = excel_workbook.sheetnames

        # Find the sheet with a name that ends with "trimmed"
        for sheet_name in sheet_names:
            if sheet_name.endswith("trimmed"):
                break  # Found the desired sheet
        else:
            # Handle the case when no matching sheet is found
            print(f"No sheet ending with 'trimmed' found in {excel_file}")
            continue

        # Load the data from the selected sheet
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name)

        # Extract accelerometer data columns (x, y, z)
        accelerometer_data = df[['x-axis (g)', 'y-axis (g)', 'z-axis (g)']].values

        for row in accelerometer_data:
            if np.all(np.isnan(row)):
                if current_trial:
                    trials.append(current_trial)
                    current_trial = []
            else:
                current_trial.append(row)

        if current_trial:
            trials.append(current_trial)

    except Exception as e:
        print(f"An error occurred while processing {excel_file}: {e}")

# Calculate the desired trial length as a multiple of 1, 2, 4, or 8
highest_trial_length = max(len(trial) for trial in trials)
desired_trial_length = highest_trial_length

while desired_trial_length % 8 != 0:
    desired_trial_length += 1

# Pad each trial with zeros to the desired length
padded_trials = []
for trial in trials:
    while len(trial) < desired_trial_length:
        trial.append(trial[-1])
    padded_trials.append(trial)

# Convert the data into a 3D tensor
tensor_data = np.array(padded_trials)
tensor_data = tensor_data.transpose(0, 2, 1)

training_seq = torch.tensor(tensor_data)

# Apply MinMax scaling to the input data and store the scalers
scaler_list = []
for i in range(training_seq.shape[0]):
    scaler = MinMaxScaler()
    reshaped_data = training_seq[i].reshape(-1, 1)
    training_seq[i] = torch.tensor(scaler.fit_transform(reshaped_data).reshape(training_seq[i].shape), dtype=torch.float)
    scaler_list.append(scaler)

training_seq = training_seq.float()

# Set up the Unet1D and GaussianDiffusion1D models
model = Unet1D(dim=64, dim_mults=(1, 2, 4, 8), channels=3)
diffusion = GaussianDiffusion1D(model, seq_length=desired_trial_length, timesteps=1000, objective='pred_v')
dataset = Dataset1D(training_seq)

# Train the model
trainer = Trainer1D(
    diffusion,
    dataset=dataset,
    train_batch_size=32,
    train_lr=8e-5,
    train_num_steps=700,         
    gradient_accumulate_every=2,    
    ema_decay=0.995,                
    amp=True,                       
)
trainer.train()

# Sample sequences from the trained model
sampled_seq = diffusion.sample(batch_size=100)
tensor_data = np.array(sampled_seq.cpu())

# Reverse the scaling and save to Excel
excel_file_path = "trial_data.xlsx"

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Remove the default sheet if it exists
if 'Sheet' in wb.sheetnames:
    std = wb['Sheet']
    wb.remove(std)

# Write each trial to a separate sheet in the workbook
for trial_number, trial_data in enumerate(tensor_data, start=1):
    sheet_title = f"Trial_{trial_number}"
    ws = wb.create_sheet(title=sheet_title)

    try:
        scaler = scaler_list[trial_number - 1]
        reshaped_data = trial_data.reshape(-1, 1)
        trial_data = scaler.inverse_transform(reshaped_data).reshape(trial_data.shape)

        reshaped_data = trial_data.T
        df = pd.DataFrame(reshaped_data, columns=[f"Feature{i}" for i in range(1, reshaped_data.shape[1] + 1)])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    except Exception as e:
        print(f"An error occurred while writing {sheet_title}: {e}")
        # You can choose to write an error message in the sheet or handle it differently.

# Save the workbook
wb.save(excel_file_path)
